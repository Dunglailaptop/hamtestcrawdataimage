from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from concurrent.futures import ThreadPoolExecutor, as_completed
from unidecode import unidecode 
from tkinter import ttk, filedialog, Tk
from datetime import datetime
from tkinter import *
from tkcalendar import DateEntry
from tkinter.filedialog import askopenfilename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import openpyxl
import pandas as pd
import math
import time
import csv
import json
import unicodedata
import re
import requests
import io
import os
from PIL import Image
import numpy as np

#folder url
urlFolder = ""
#file excel url 
urlFileExcel = ""
#ngày thiết lập lấy
dateSelect = ""


def login():
    global dateSelect
    try:
        #khai báo thông số     
        chromedriver_path = "chromedriver.exe"  # Ensure this path is correct
        login_url = "https://vnur.vn/"
        area_data_url = ""
        username = "quyen.ngoq"
        password = "74777477"
        area_data_url = "http://192.168.0.65:8180/#menu=131&action=111"  
            

        if not os.path.isfile(chromedriver_path):
            raise ValueError(f"The path is not a valid file: {chromedriver_path}")
        
        print(f"Using chromedriver at: {chromedriver_path}")
        # # # Initialize ChromeDriver
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-plugins")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")

        #     # Thêm các tùy chọn để giảm tải CPU và bộ nhớ
        # options.add_argument("--no-zygote")
        # options.add_argument("--single-process")
        # options.add_argument("--disable-setuid-sandbox")
        # options.add_argument("--ignore-certificate-errors")
        # options.add_argument("--disable-accelerated-2d-canvas")
        # options.add_argument("--disable-gpu-sandbox")
        service = Service(chromedriver_path)
        driver = webdriver.Chrome(service=service,options=options)
        
        
       

        # Open the website
        driver.get(login_url)
        driver.maximize_window()

        # # Wait for the page to load
        # time.sleep(1)

        # # Find and enter username
        # username_input = driver.find_element(By.ID, "txtUsername")
        # username_input.send_keys(username)

        # # Find and enter password
        # password_input = driver.find_element(By.ID, "txtPassword")
        # password_input.send_keys(password)

        # # Find and click the login button
        # login_button = driver.find_element(By.ID, "btnLogin")
        # login_button.click()

        # # Wait for login to complete
        # time.sleep(5)

        # Click save button
        # save = driver.find_element(By.ID, "btnSave")
        # save.click()
        # time.sleep(3)
        
        # chọn phân trang tính toán tổng số page
        # driver.get(area_data_url)
        # time.sleep(2)
        # set_date2(driver, "dbFrom",dateSelect)
        # set_date2(driver, "dbTo", dateSelect)
    except Exception as e:
        print(f"Lỗi hàm login: {e}")

    return driver

def main():
    try:
        # Đăng nhập lần đầu để lấy giá trị total
        driver = login()
        time.sleep(3)
        #hàm lấy tổng total ==> trả tổng page, và tổng số lượng phần tử trong ngày 01/01/2023
        total = 300  # Hàm này cần được định nghĩa để lấy giá trị total sau khi đăng nhập
        driver.quit()
        #hàm lấy dữ liệu file excel ==> trả về thằng phẩn tử có số thứ tự cuối cùng 
        #gán cho hàm numberget 
        #trường hợp 2 là file excel rỗng ===> tính từ thằng có phần tử số stt 1 và page 1


        totalfinal = total  # Sử dụng total làm giới hạn cho vòng lặp
        numberget = 0
        success = False
        driver = None
        
        while numberget < totalfinal:
            print(f"====đang chạy: {numberget}====")
            
            if numberget % 4 == 0:  # Đăng nhập lại sau mỗi 4 lần lặp
                if driver:
                    driver.quit()  # Đóng phiên đăng nhập cũ nếu có
                driver = login()
                time.sleep(3)
            
            # Thực hiện các hành động khác ở đây
            # Ví dụ: xử lý dữ liệu, thao tác trên trang web, v.v.
            
            numberget += 1
        
        if driver:
            driver.quit()  # Đảm bảo đóng phiên cuối cùng
        success = True
        
    except Exception as e:
        print(f"Lỗi hàm main: {e}")
        if driver:
            driver.quit()  # Đảm bảo đóng driver nếu có lỗi xảy ra

    return success



#chọn folder để lưu
def select_folder():
    global urlFolder
    folder_path = filedialog.askdirectory(
        title="Select a folder"
    )
    if folder_path:
        print(folder_path)
        urlFolder = folder_path
        main()
    else:
        print("No folder selected.")

#lấy file excel
def select_file_Excel():
   global urlFileExcel
   file_path = filedialog.askopenfilename(
       title="select file excel",
       filetypes=[("Excel files","*.xlsx;*.xls")]
   )
   print(f"chọn thành công file excel: {file_path}")
   if file_path:
      urlFileExcel = file_path
      get_data_file_excel(file_path)
   else:
      print("không có file excel được chọn")
#đọc dữ liệu file excel lấy page số trang đang tới vị trí hiện tại và số stt của phần tử lấy thông tin ngày lấy dữ liệu     
def get_data_file_excel(file_path):
    global dateSelect
    if not file_path:
        print("Không có file nào được chọn.")
        return None
    # tìm tên file 
    file_name = os.path.basename(file_path)
    name_part, extension = os.path.splitext(file_name)
    try:
        date = datetime.strptime(name_part,"%d-%m-%Y")
        formatted_date = date.strftime("%d/%m/%Y")
        dateSelect = formatted_date
        
        # Đọc workbook
        # workbook = openpyxl.load_workbook(file_path)
        # sheet = workbook.active
        # data = []
        # for row in sheet.iter_rows(values_only=True):
        #     data.append(row)
        print(formatted_date)
    except ValueError:
        print(name_part)


     

root = Tk()
root.title("Tkinter ComboBox Example")

#so 
numberget = [0]
# Đặt kích thước cho cửa sổ
window_width = 400
window_height = 200

# Lấy kích thước màn hình
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Tính toán vị trí để cửa sổ nằm giữa màn hình
position_x = (screen_width // 2) - (window_width // 2)
position_y = (screen_height // 2) - (window_height // 2)

# Đặt kích thước và vị trí cho cửa sổ
root.geometry(f'{window_width}x{window_height}+{position_x}+{position_y}')


date1 = ''
date2 = ''
    # Create a label to display instructions
label = ttk.Label(root, text="SIÊU PHẦN MỀM CRAW ĐÁ DỮ LIỆU ẢNH")
label.pack(pady=10)
# Create a label to display instructions
label = ttk.Label(root, text="Choose an option:")
label.pack(pady=10)


# #get data json
file_button = ttk.Button(root, text="Select csv Excel", command=select_file_Excel, width=30)  # Corrected here
file_button.pack(pady=10)



button = ttk.Button(root, text="Get Data", command=select_folder, width=10)
button.pack(pady=10)

# Start the Tkinter event loop
root.mainloop()
